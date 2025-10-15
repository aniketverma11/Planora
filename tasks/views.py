from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.http import HttpResponse
from .models import Task, TaskDocument
from .serializers import (
    TaskSerializer,
    TaskCreateUpdateSerializer,
    TaskDocumentSerializer,
)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from datetime import datetime
from .critical_path import calculate_critical_path

class TaskViewSet(viewsets.ModelViewSet):
    queryset = Task.objects.all()
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return TaskCreateUpdateSerializer
        return TaskSerializer

    def get_queryset(self):
        queryset = Task.objects.all()
        
        # Filter by project if project_id is provided
        project_id = self.request.query_params.get('project_id', None)
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        
        # Filter to show only main tasks (not subtasks) unless showing all
        if self.action not in ['all_tasks']:
            queryset = queryset.filter(parent_task__isnull=True)
        
        return queryset
    
    @action(detail=True, methods=['get'])
    def subtasks(self, request, pk=None):
        """Get all subtasks for a specific task"""
        task = self.get_object()
        subtasks = task.subtasks.all()
        serializer = TaskSerializer(subtasks, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def all_tasks(self, request):
        """Get all tasks including subtasks for Gantt chart"""
        # Get base queryset (without parent_task filter)
        queryset = Task.objects.all()
        
        # Filter by project if project_id is provided
        project_id = self.request.query_params.get('project_id', None)
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        
        serializer = TaskSerializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def upload_document(self, request, pk=None):
        """Upload a document for a task"""
        task = self.get_object()
        file = request.FILES.get('file')
        
        if not file:
            return Response(
                {'error': 'No file provided'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Create document
        document = TaskDocument(
            task=task,
            file=file,
            file_name=file.name,
            file_size=file.size,
            file_type=file.content_type,
            uploaded_by=request.user
        )
        document.save()
        
        serializer = TaskDocumentSerializer(document)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['delete'], url_path='delete_document/(?P<document_id>[^/.]+)')
    def delete_document(self, request, pk=None, document_id=None):
        """Delete a document from a task"""
        task = self.get_object()
        
        try:
            document = TaskDocument.objects.get(id=document_id, task=task)
            document.file.delete()  # Delete the actual file
            document.delete()  # Delete the database record
            return Response(status=status.HTTP_204_NO_CONTENT)
        except TaskDocument.DoesNotExist:
            return Response(
                {'error': 'Document not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=True, methods=['get'])
    def documents(self, request, pk=None):
        """Get all documents for a task"""
        task = self.get_object()
        documents = task.documents.all()
        serializer = TaskDocumentSerializer(documents, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export tasks to Excel file"""
        # Get project_id from query params
        project_id = request.query_params.get('project_id')
        
        if not project_id:
            return Response(
                {'error': 'project_id is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Filter tasks by project
        tasks = Task.objects.filter(project_id=project_id).order_by('id')
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Tasks"
        
        # Define headers
        headers = [
            'ID', 'Title', 'Description', 'Status', 'Priority',
            'Start Date', 'Due Date', 'Duration (days)', 'Progress (%)',
            'Parent Task ID', 'Assignee Email', 'Dependencies (IDs)'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data
        for row_num, task in enumerate(tasks, 2):
            ws.cell(row=row_num, column=1, value=task.id)
            ws.cell(row=row_num, column=2, value=task.title)
            ws.cell(row=row_num, column=3, value=task.description or '')
            ws.cell(row=row_num, column=4, value=task.status)
            ws.cell(row=row_num, column=5, value=task.priority)
            ws.cell(row=row_num, column=6, value=task.start_date.strftime('%Y-%m-%d') if task.start_date else '')
            ws.cell(row=row_num, column=7, value=task.due_date.strftime('%Y-%m-%d') if task.due_date else '')
            ws.cell(row=row_num, column=8, value=task.duration or '')
            ws.cell(row=row_num, column=9, value=task.progress or 0)
            ws.cell(row=row_num, column=10, value=task.parent_task_id or '')
            ws.cell(row=row_num, column=11, value=task.assignee.email if task.assignee else '')
            
            # Dependencies as comma-separated IDs
            dep_ids = ','.join(str(dep.id) for dep in task.dependencies.all()) if task.dependencies.exists() else ''
            ws.cell(row=row_num, column=12, value=dep_ids)
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create response
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=tasks_project_{project_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
    
    @action(detail=False, methods=['get'])
    def download_sample(self, request):
        """Download sample Excel template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Tasks"
        
        # Define headers
        headers = [
            'Title', 'Description', 'Status', 'Priority',
            'Start Date', 'Due Date', 'Duration (days)', 'Progress (%)',
            'Parent Task ID', 'Assignee Email', 'Dependencies (IDs)'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Add sample data with emails instead of usernames
        sample_data = [
            ['Sample Task 1', 'This is a sample task', 'To Do', 'High', '2025-01-01', '2025-01-10', 10, 0, '', 'admin@example.com', ''],
            ['Sample Task 2', 'Another sample task', 'In Progress', 'Medium', '2025-01-05', '2025-01-15', 10, 50, '', 'user@example.com', ''],
        ]
        
        for row_num, data in enumerate(sample_data, 2):
            for col_num, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions")
        instructions = [
            "Instructions for importing tasks:",
            "",
            "1. Fill in the task details in the 'Tasks' sheet",
            "2. Required fields: Title, Status, Priority",
            "3. Status options: To Do, In Progress, Done",
            "4. Priority options: Low, Medium, High",
            "5. Date format: YYYY-MM-DD (e.g., 2025-01-01)",
            "6. Duration: Number of days",
            "7. Progress: Number between 0-100",
            "8. Parent Task ID: Leave empty for main tasks, enter task ID for subtasks",
            "9. Assignee Email: Must match existing user email (e.g., admin@example.com)",
            "10. Dependencies: Comma-separated task IDs (e.g., 1,2,3)",
            "",
            "Note: Do not modify the header row in the Tasks sheet"
        ]
        
        for row_num, instruction in enumerate(instructions, 1):
            ws_instructions.cell(row=row_num, column=1, value=instruction)
        
        ws_instructions.column_dimensions['A'].width = 80
        
        # Adjust column widths in Tasks sheet
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create response
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=task_import_template.xlsx'
        
        return response
    
    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        """Import tasks from Excel file"""
        project_id = request.data.get('project_id')
        
        if not project_id:
            return Response(
                {'error': 'project_id is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES.get('file')
        if not file:
            return Response(
                {'error': 'No file provided'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Load workbook
            wb = load_workbook(file)
            ws = wb.active
            
            # Validate headers
            expected_headers = [
                'Title', 'Description', 'Status', 'Priority',
                'Start Date', 'Due Date', 'Duration (days)', 'Progress (%)',
                'Parent Task ID', 'Assignee Email', 'Dependencies (IDs)'
            ]
            
            actual_headers = [cell.value for cell in ws[1]]
            
            # Skip validation if first column is 'ID' (exported file)
            if actual_headers[0] == 'ID':
                actual_headers = actual_headers[1:]  # Remove ID column
                start_col = 2
            else:
                start_col = 1
            
            created_tasks = []
            errors = []
            
            # Process each row
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
                try:
                    # Get values based on start column
                    if start_col == 2:
                        values = [cell.value for cell in row[1:]]  # Skip ID column
                    else:
                        values = [cell.value for cell in row]
                    
                    if not values[0]:  # Skip if title is empty
                        continue
                    
                    # Parse data
                    title = str(values[0]).strip()
                    description = str(values[1]).strip() if values[1] else ''
                    status_val = str(values[2]).strip().lower() if values[2] else 'pending'
                    priority = str(values[3]).strip().lower() if values[3] else 'medium'
                    
                    # Parse dates
                    start_date = None
                    if values[4]:
                        if isinstance(values[4], datetime):
                            start_date = values[4].date()
                        elif isinstance(values[4], str):
                            try:
                                start_date = datetime.strptime(values[4], '%Y-%m-%d').date()
                            except:
                                pass
                    
                    due_date = None
                    if values[5]:
                        if isinstance(values[5], datetime):
                            due_date = values[5].date()
                        elif isinstance(values[5], str):
                            try:
                                due_date = datetime.strptime(values[5], '%Y-%m-%d').date()
                            except:
                                pass
                    
                    duration = int(values[6]) if values[6] else None
                    progress = int(values[7]) if values[7] else 0
                    parent_task_id = int(values[8]) if values[8] else None
                    assignee_email = str(values[9]).strip() if values[9] else None
                    dependency_ids = str(values[10]).strip() if values[10] else ''
                    
                    # Get assignee by email (more robust than username)
                    assignee = None
                    if assignee_email:
                        from users.models import CustomUser
                        try:
                            assignee = CustomUser.objects.get(email=assignee_email)
                        except CustomUser.DoesNotExist:
                            errors.append(f"Row {row_num}: User with email '{assignee_email}' not found")
                            # Don't skip - create task without assignee
                            assignee = None
                        except CustomUser.MultipleObjectsReturned:
                            # If multiple users have same email (shouldn't happen), take the first
                            assignee = CustomUser.objects.filter(email=assignee_email).first()
                            errors.append(f"Row {row_num}: Warning - Multiple users found with email '{assignee_email}', using first match")
                    
                    # Create task
                    task = Task.objects.create(
                        title=title,
                        description=description,
                        status=status_val,
                        priority=priority,
                        start_date=start_date,
                        due_date=due_date,
                        duration=duration,
                        progress=progress,
                        parent_task_id=parent_task_id,
                        assignee=assignee,
                        project_id=project_id
                    )
                    
                    # Add dependencies
                    if dependency_ids:
                        dep_ids = [int(dep_id.strip()) for dep_id in dependency_ids.split(',') if dep_id.strip().isdigit()]
                        for dep_id in dep_ids:
                            try:
                                dep_task = Task.objects.get(id=dep_id, project_id=project_id)
                                task.dependencies.add(dep_task)
                            except Task.DoesNotExist:
                                errors.append(f"Row {row_num}: Dependency task {dep_id} not found")
                    
                    created_tasks.append(task.id)
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
            
            return Response({
                'success': True,
                'created_count': len(created_tasks),
                'created_task_ids': created_tasks,
                'errors': errors
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response(
                {'error': f'Failed to process Excel file: {str(e)}'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['get'])
    def critical_path(self, request):
        """Get critical path analysis for a project"""
        project_id = request.query_params.get('project_id')
        
        if not project_id:
            return Response(
                {'error': 'project_id is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Get all tasks for the project
            tasks = Task.objects.filter(project_id=project_id).prefetch_related('dependencies', 'dependents')
            
            if not tasks.exists():
                return Response({
                    'critical_tasks': [],
                    'critical_paths': [],
                    'project_duration': 0,
                    'total_tasks': 0,
                    'critical_tasks_count': 0,
                    'message': 'No tasks found for this project'
                })
            
            # Calculate critical path
            result = calculate_critical_path(tasks)
            
            # Serialize critical tasks
            critical_tasks_data = []
            for task in result['critical_tasks']:
                critical_tasks_data.append({
                    'id': task.id,
                    'title': task.title,
                    'description': task.description or '',
                    'duration': task.duration,
                    'early_start': task.early_start_day,
                    'early_finish': task.early_finish_day,
                    'late_start': task.late_start_day,
                    'late_finish': task.late_finish_day,
                    'total_float': task.total_float,
                    'is_critical': task.is_critical,
                    'status': task.status,
                    'progress': task.progress,
                    'assignee_username': task.assignee.username if task.assignee else None
                })
            
            return Response({
                'critical_tasks': critical_tasks_data,
                'critical_paths': result['critical_paths'],
                'project_duration': result['project_duration'],
                'earliest_completion': result['earliest_completion'],
                'latest_completion': result['latest_completion'],
                'total_tasks': result['total_tasks'],
                'critical_tasks_count': result['critical_tasks_count'],
                'risk_level': result['risk_level']
            })
            
        except ValueError as e:
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'error': f'Failed to calculate critical path: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'])
    def calculate_critical_path(self, request):
        """Calculate and save critical path data for a project"""
        project_id = request.data.get('project_id')
        
        if not project_id:
            return Response(
                {'error': 'project_id is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Get all tasks for the project
            tasks = Task.objects.filter(project_id=project_id).prefetch_related('dependencies', 'dependents')
            
            if not tasks.exists():
                return Response({
                    'success': False,
                    'message': 'No tasks found for this project'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Calculate critical path
            result = calculate_critical_path(tasks)
            
            # Save the calculated values to database
            for task in tasks:
                task.save(update_fields=[
                    'early_start_day', 'early_finish_day',
                    'late_start_day', 'late_finish_day',
                    'total_float', 'is_critical'
                ])
            
            return Response({
                'success': True,
                'message': 'Critical path calculated and saved successfully',
                'project_duration': result['project_duration'],
                'critical_tasks_count': result['critical_tasks_count'],
                'risk_level': result['risk_level']
            }, status=status.HTTP_200_OK)
            
        except ValueError as e:
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'error': f'Failed to calculate critical path: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def float_analysis(self, request):
        """Get float/slack analysis for all tasks in a project"""
        project_id = request.query_params.get('project_id')
        
        if not project_id:
            return Response(
                {'error': 'project_id is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Get all tasks for the project
            tasks = Task.objects.filter(project_id=project_id).prefetch_related('dependencies', 'assignee')
            
            if not tasks.exists():
                return Response({
                    'critical': [],
                    'near_critical': [],
                    'normal': [],
                    'message': 'No tasks found for this project'
                })
            
            # Calculate critical path to ensure data is up to date
            calculate_critical_path(tasks)
            
            # Categorize tasks by float
            critical_tasks = []
            near_critical_tasks = []
            normal_tasks = []
            
            for task in tasks:
                task_data = {
                    'id': task.id,
                    'title': task.title,
                    'description': task.description or '',
                    'duration': task.duration,
                    'total_float': task.total_float,
                    'early_start': task.early_start_day,
                    'early_finish': task.early_finish_day,
                    'late_start': task.late_start_day,
                    'late_finish': task.late_finish_day,
                    'status': task.status,
                    'progress': task.progress,
                    'assignee_username': task.assignee.username if task.assignee else None
                }
                
                if task.total_float == 0:
                    critical_tasks.append(task_data)
                elif task.total_float <= 2:
                    near_critical_tasks.append(task_data)
                else:
                    normal_tasks.append(task_data)
            
            return Response({
                'critical': critical_tasks,
                'near_critical': near_critical_tasks,
                'normal': normal_tasks,
                'summary': {
                    'critical': len(critical_tasks),
                    'near_critical': len(near_critical_tasks),
                    'normal': len(normal_tasks),
                    'total': len(tasks)
                }
            })
            
        except Exception as e:
            return Response(
                {'error': f'Failed to analyze float: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
