from rest_framework import generics, status, viewsets
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.decorators import action
from .serializers import UserSerializer, UserListSerializer, AvailablePermissionsSerializer, CustomTokenObtainPairSerializer
from .models import CustomUser
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework import permissions
from .google_auth import verify_google_token
from .microsoft_auth import verify_microsoft_token
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import Permission
from django.contrib.contenttypes.models import ContentType
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime
import secrets


class IsAdminUser(permissions.BasePermission):
    """
    Custom permission to only allow admin users to access user management
    """
    def has_permission(self, request, view):
        return request.user and request.user.is_authenticated and request.user.designation == 'admin'

class SignUpView(generics.CreateAPIView):
    queryset = CustomUser.objects.all()
    serializer_class = UserSerializer

class LoginView(TokenObtainPairView):
    permission_classes = (permissions.AllowAny,)
    serializer_class = CustomTokenObtainPairSerializer

class UserListView(generics.ListAPIView):
    queryset = CustomUser.objects.all()
    serializer_class = UserSerializer
    permission_classes = (permissions.IsAuthenticated,)


class GoogleAuthView(APIView):
    """
    Handle Google OAuth authentication
    """
    permission_classes = (permissions.AllowAny,)

    def post(self, request):
        try:
            token = request.data.get('token')
            if not token:
                return Response(
                    {'error': 'Token is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Verify the Google token and get user info
            user_info = verify_google_token(token)

            # Check if user exists with this Google ID
            user = CustomUser.objects.filter(google_id=user_info['google_id']).first()

            if user:
                # User exists, log them in
                if not user.is_active:
                    return Response(
                        {'error': 'User account is disabled'},
                        status=status.HTTP_403_FORBIDDEN
                    )
            else:
                # Check if user exists with this email (from regular signup)
                user = CustomUser.objects.filter(email=user_info['email']).first()
                
                if user:
                    # Link Google account to existing user
                    user.google_id = user_info['google_id']
                    user.auth_provider = 'google'
                    user.profile_picture = user_info.get('picture', '')
                    user.save()
                else:
                    # Create new user
                    username = user_info['email'].split('@')[0]
                    base_username = username
                    counter = 1
                    
                    # Ensure unique username
                    while CustomUser.objects.filter(username=username).exists():
                        username = f"{base_username}{counter}"
                        counter += 1

                    user = CustomUser.objects.create(
                        username=username,
                        email=user_info['email'],
                        first_name=user_info.get('first_name', ''),
                        last_name=user_info.get('last_name', ''),
                        google_id=user_info['google_id'],
                        profile_picture=user_info.get('picture', ''),
                        auth_provider='google',
                        password=make_password(secrets.token_urlsafe(32)),  # Random password
                    )

            # Generate JWT tokens
            refresh = RefreshToken.for_user(user)
            
            return Response({
                'access': str(refresh.access_token),
                'refresh': str(refresh),
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'profile_picture': user.profile_picture,
                    'auth_provider': user.auth_provider,
                    'designation': user.designation,
                }
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )


class MicrosoftAuthView(APIView):
    """
    Handle Microsoft OAuth authentication
    """
    permission_classes = (permissions.AllowAny,)

    def post(self, request):
        try:
            token = request.data.get('token')
            if not token:
                return Response(
                    {'error': 'Token is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Verify the Microsoft token and get user info
            user_info = verify_microsoft_token(token)

            # Check if user exists with this Microsoft ID
            user = CustomUser.objects.filter(microsoft_id=user_info['microsoft_id']).first()

            if user:
                # User exists, log them in
                if not user.is_active:
                    return Response(
                        {'error': 'User account is disabled'},
                        status=status.HTTP_403_FORBIDDEN
                    )
            else:
                # Check if user exists with this email (from regular signup)
                user = CustomUser.objects.filter(email=user_info['email']).first()
                
                if user:
                    # Link Microsoft account to existing user
                    user.microsoft_id = user_info['microsoft_id']
                    user.auth_provider = 'microsoft'
                    user.profile_picture = user_info.get('profile_picture', '')
                    user.save()
                else:
                    # Create new user
                    username = user_info['email'].split('@')[0]
                    base_username = username
                    counter = 1
                    
                    # Ensure unique username
                    while CustomUser.objects.filter(username=username).exists():
                        username = f"{base_username}{counter}"
                        counter += 1

                    user = CustomUser.objects.create(
                        username=username,
                        email=user_info['email'],
                        first_name=user_info.get('given_name', ''),
                        last_name=user_info.get('family_name', ''),
                        microsoft_id=user_info['microsoft_id'],
                        profile_picture=user_info.get('profile_picture', ''),
                        auth_provider='microsoft',
                        password=make_password(secrets.token_urlsafe(32)),  # Random password
                    )

            # Generate JWT tokens
            refresh = RefreshToken.for_user(user)
            
            return Response({
                'access': str(refresh.access_token),
                'refresh': str(refresh),
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'profile_picture': user.profile_picture,
                    'auth_provider': user.auth_provider,
                    'designation': user.designation,
                }
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )



class UserManagementViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing users - only accessible to admin users
    """
    queryset = CustomUser.objects.all()
    permission_classes = [IsAdminUser]
    
    def get_serializer_class(self):
        if self.action == "list":
            return UserListSerializer
        return UserSerializer
    
    def create(self, request, *args, **kwargs):
        """Create user and send verification email"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Create user without password (will be set via email verification)
        user = serializer.save()
        user.set_unusable_password()  # User must set password via email
        user.email_verified = False
        user.save()
        
        # Generate verification token
        token = user.generate_verification_token()
        
        # Send verification email
        email_sent = False
        email_error = None
        try:
            from utils.email_service import email_service
            from utils.email_templates.templates import verification_email_template
            from django.conf import settings
            
            verification_link = f"{settings.FRONTEND_URL}/verify-email?token={token}"
            html_content = verification_email_template(user, verification_link)
            
            print(f"📧 Attempting to send verification email to {user.email}")
            print(f"🔗 Verification link: {verification_link}")
            
            email_sent = email_service.send_email(
                to_email=user.email,
                subject='Welcome to Task Management System - Verify Your Email',
                html_content=html_content
            )
            
            if email_sent:
                print(f"✅ Verification email sent successfully to {user.email}")
            else:
                print(f"❌ Failed to send verification email to {user.email}")
                email_error = "Email service returned False"
                
        except Exception as e:
            email_error = str(e)
            print(f"❌ Error sending verification email: {email_error}")
            import traceback
            traceback.print_exc()
        
        headers = self.get_success_headers(serializer.data)
        return Response(
            {
                'message': 'User created successfully. Verification email sent.' if email_sent else f'User created but email failed: {email_error}',
                'user': serializer.data,
                'email_sent': email_sent,
                'verification_link': f"{settings.FRONTEND_URL}/verify-email?token={token}" if not email_sent else None  # Include link if email fails
            },
            status=status.HTTP_201_CREATED,
            headers=headers
        )
    
    def get_queryset(self):
        """Allow filtering and searching users"""
        queryset = CustomUser.objects.all().order_by("-date_joined")
        
        # Filter by designation
        designation = self.request.query_params.get("designation", None)
        if designation:
            queryset = queryset.filter(designation=designation)
        
        # Filter by active status
        is_active = self.request.query_params.get("is_active", None)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == "true")
        
        # Search by username or email
        search = self.request.query_params.get("search", None)
        if search:
            queryset = queryset.filter(
                username__icontains=search
            ) | queryset.filter(
                email__icontains=search
            ) | queryset.filter(
                first_name__icontains=search
            ) | queryset.filter(
                last_name__icontains=search
            )
        
        return queryset
    
    @action(detail=False, methods=["get"])
    def available_permissions(self, request):
        """
        Get all available permissions for tasks, projects, and users
        """
        # Get content types for our models
        apps_to_include = ["tasks", "project", "users"]
        
        permissions_data = []
        
        for app_label in apps_to_include:
            content_types = ContentType.objects.filter(app_label=app_label)
            
            for ct in content_types:
                permissions = Permission.objects.filter(content_type=ct)
                if permissions.exists():
                    permissions_data.append({
                        "app_label": app_label,
                        "model": ct.model,
                        "model_name": ct.model_class()._meta.verbose_name if ct.model_class() else ct.model,
                        "permissions": [
                            {
                                "id": p.id,
                                "name": p.name,
                                "codename": p.codename,
                                "content_type": p.content_type_id
                            }
                            for p in permissions
                        ]
                    })
        
        return Response(permissions_data)
    
    @action(detail=False, methods=["get"])
    def current_user(self, request):
        """
        Get current logged-in user with designation
        """
        serializer = UserSerializer(request.user)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'], url_path='import-excel')
    def import_excel(self, request):
        """
        Import users from Excel file
        Expected columns: Username, First Name, Last Name, Email, Password (optional)
        """
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file uploaded'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        excel_file = request.FILES['file']
        
        try:
            # Load the workbook
            wb = load_workbook(excel_file)
            ws = wb.active
            
            # Read headers
            headers = [cell.value for cell in ws[1]]
            
            # Validate headers
            required_headers = ['Username', 'First Name', 'Last Name', 'Email']
            for header in required_headers:
                if header not in headers:
                    return Response(
                        {'error': f'Missing required column: {header}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            # Get column indices
            username_idx = headers.index('Username')
            first_name_idx = headers.index('First Name')
            last_name_idx = headers.index('Last Name')
            email_idx = headers.index('Email')
            password_idx = headers.index('Password') if 'Password' in headers else None
            
            created_count = 0
            skipped_count = 0
            errors = []
            
            # Process each row (skip header)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Skip empty rows
                    continue
                
                username = row[username_idx]
                first_name = row[first_name_idx]
                last_name = row[last_name_idx]
                email = row[email_idx]
                password = row[password_idx] if password_idx is not None and row[password_idx] else None
                
                # Validate required fields
                if not username or not email:
                    errors.append(f'Row {row_num}: Username and Email are required')
                    skipped_count += 1
                    continue
                
                # Check if user already exists
                if CustomUser.objects.filter(username=username).exists():
                    errors.append(f'Row {row_num}: User "{username}" already exists')
                    skipped_count += 1
                    continue
                
                if CustomUser.objects.filter(email=email).exists():
                    errors.append(f'Row {row_num}: Email "{email}" already exists')
                    skipped_count += 1
                    continue
                
                try:
                    # Create user
                    user = CustomUser.objects.create(
                        username=username,
                        email=email,
                        first_name=first_name or '',
                        last_name=last_name or '',
                        designation='user',  # Default designation
                        email_verified=False
                    )
                    
                    # If password provided, set it and mark as verified
                    if password:
                        user.set_password(password)
                        user.email_verified = True
                        user.save()
                    else:
                        # No password provided, send verification email
                        user.set_unusable_password()
                        user.save()
                        
                        # Generate verification token
                        token = user.generate_verification_token()
                        
                        # Send verification email
                        try:
                            from utils.email_service import email_service
                            from utils.email_templates.templates import verification_email_template
                            from django.conf import settings
                            
                            verification_link = f"{settings.FRONTEND_URL}/verify-email?token={token}"
                            html_content = verification_email_template(user, verification_link)
                            
                            email_service.send_email(
                                to_email=user.email,
                                subject='Welcome to Task Management System - Verify Your Email',
                                html_content=html_content
                            )
                        except Exception as email_error:
                            print(f"❌ Error sending verification email to {user.email}: {str(email_error)}")
                    
                    created_count += 1
                    
                except Exception as e:
                    errors.append(f'Row {row_num}: {str(e)}')
                    skipped_count += 1
            
            return Response({
                'message': 'Import completed',
                'created_count': created_count,
                'skipped_count': skipped_count,
                'errors': errors[:20] if errors else []  # Limit errors to first 20
            })
            
        except Exception as e:
            return Response(
                {'error': f'Failed to process Excel file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """
        Export all users to Excel file
        """
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        
        # Define headers
        headers = ['Username', 'First Name', 'Last Name', 'Email', 'Designation', 'Date Joined', 'Is Active']
        
        # Style for headers
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Get all users
        users = self.get_queryset()
        
        # Write data
        for row, user in enumerate(users, start=2):
            ws.cell(row=row, column=1, value=user.username)
            ws.cell(row=row, column=2, value=user.first_name)
            ws.cell(row=row, column=3, value=user.last_name)
            ws.cell(row=row, column=4, value=user.email)
            ws.cell(row=row, column=5, value=user.designation)
            ws.cell(row=row, column=6, value=user.date_joined.strftime('%Y-%m-%d %H:%M:%S') if user.date_joined else '')
            ws.cell(row=row, column=7, value='Yes' if user.is_active else 'No')
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create response
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=users_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
    
    @action(detail=False, methods=['get'], url_path='download-sample')
    def download_sample(self, request):
        """
        Download sample Excel template for user import
        """
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Users Template"
        
        # Define headers
        headers = ['Username', 'First Name', 'Last Name', 'Email', 'Password']
        
        # Style for headers
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add sample data
        sample_data = [
            ['john.doe', 'John', 'Doe', 'john.doe@example.com', 'password123'],
            ['jane.smith', 'Jane', 'Smith', 'jane.smith@example.com', ''],  # Empty password will be auto-generated
            ['bob.wilson', 'Bob', 'Wilson', 'bob.wilson@example.com', 'secure456'],
        ]
        
        for row_num, data in enumerate(sample_data, start=2):
            for col, value in enumerate(data, start=1):
                ws.cell(row=row_num, column=col, value=value)
        
        # Add instructions
        ws.cell(row=6, column=1, value='Instructions:')
        ws.cell(row=7, column=1, value='1. Fill in Username, First Name, Last Name, and Email (required)')
        ws.cell(row=8, column=1, value='2. Password is optional - if empty, a random password will be generated')
        ws.cell(row=9, column=1, value='3. Remove sample data before importing')
        ws.cell(row=10, column=1, value='4. Save and upload this file')
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 25
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create response
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=sample_user_import_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
    
    @action(detail=True, methods=['post'], url_path='resend-verification')
    def resend_verification_email(self, request, pk=None):
        """
        Resend verification email to a specific user
        """
        user = self.get_object()
        
        # Check if user is already verified
        if user.email_verified:
            return Response(
                {'error': 'User email is already verified.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if user has a usable password
        if user.has_usable_password():
            return Response(
                {'error': 'User already has a password set.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            from utils.email_service import email_service
            from utils.email_templates.templates import verification_email_template
            from django.conf import settings
            
            # Generate new verification token
            token = user.generate_verification_token()
            
            # Create verification link
            verification_link = f"{settings.FRONTEND_URL}/verify-email?token={token}"
            html_content = verification_email_template(user, verification_link)
            
            # Send email
            email_sent = email_service.send_email(
                to_email=user.email,
                subject='Welcome to Task Management System - Verify Your Email',
                html_content=html_content
            )
            
            if email_sent:
                return Response({
                    'message': f'Verification email sent successfully to {user.email}',
                    'email': user.email
                })
            else:
                return Response(
                    {'error': 'Failed to send email. Please try again.'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
                
        except Exception as e:
            return Response(
                {'error': f'Failed to send verification email: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CurrentUserView(APIView):
    """
    Get current user information
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        serializer = UserSerializer(request.user)
        return Response(serializer.data)


class VerifyEmailView(APIView):
    """
    Verify email and allow password setup
    """
    permission_classes = (permissions.AllowAny,)
    
    def post(self, request):
        try:
            token = request.data.get('token')
            password = request.data.get('password')
            
            if not token:
                return Response(
                    {'error': 'Verification token is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Find user with this token
            user = CustomUser.objects.filter(verification_token=token).first()
            
            if not user:
                return Response(
                    {'error': 'Invalid verification token'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Check if token is still valid
            if not user.is_verification_token_valid():
                return Response(
                    {'error': 'Verification token has expired. Please request a new one.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # If password is provided, set it
            if password:
                user.set_password(password)
            
            # Mark email as verified
            user.email_verified = True
            user.verification_token = None
            user.verification_token_created_at = None
            user.save()
            
            # Send success confirmation email
            try:
                from utils.email_service import email_service
                from utils.email_templates.templates import email_verification_success_template
                from django.conf import settings
                
                login_url = f"{settings.FRONTEND_URL}/login"
                html_content = email_verification_success_template(user, login_url)
                
                email_service.send_email(
                    to_email=user.email,
                    subject='Welcome! Your Account is Now Active',
                    html_content=html_content
                )
                print(f"✅ Verification success email sent to {user.email}")
            except Exception as email_error:
                print(f"❌ Failed to send verification success email: {str(email_error)}")
                # Don't fail the verification if email fails
            
            return Response({
                'message': 'Email verified successfully',
                'username': user.username,
                'email': user.email
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ResendVerificationEmailView(APIView):
    """
    Resend verification email if expired or lost
    """
    permission_classes = (permissions.AllowAny,)
    
    def post(self, request):
        try:
            email = request.data.get('email')
            
            if not email:
                return Response(
                    {'error': 'Email is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Find user with this email
            user = CustomUser.objects.filter(email=email).first()
            
            if not user:
                return Response(
                    {'error': 'No user found with this email'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Check if already verified
            if user.email_verified:
                return Response(
                    {'error': 'Email is already verified. You can log in.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Generate new verification token
            token = user.generate_verification_token()
            
            # Send verification email
            from utils.email_service import email_service
            from utils.email_templates.templates import verification_email_template
            from django.conf import settings
            
            verification_link = f"{settings.FRONTEND_URL}/verify-email?token={token}"
            html_content = verification_email_template(user, verification_link)
            
            success = email_service.send_email(
                to_email=user.email,
                subject='Verify Your Email - Task Management System',
                html_content=html_content
            )
            
            if success:
                return Response({
                    'message': 'Verification email sent successfully. Please check your inbox.'
                }, status=status.HTTP_200_OK)
            else:
                return Response(
                    {'error': 'Failed to send verification email. Please try again later.'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
                
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CheckVerificationTokenView(APIView):
    """
    Check if verification token is valid
    """
    permission_classes = (permissions.AllowAny,)
    
    def get(self, request):
        try:
            token = request.query_params.get('token')
            
            if not token:
                return Response(
                    {'error': 'Token is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            user = CustomUser.objects.filter(verification_token=token).first()
            
            if not user:
                return Response({
                    'valid': False,
                    'message': 'Invalid token'
                }, status=status.HTTP_200_OK)
            
            is_valid = user.is_verification_token_valid()
            
            return Response({
                'valid': is_valid,
                'message': 'Token is valid' if is_valid else 'Token has expired',
                'email': user.email if is_valid else None,
                'username': user.username if is_valid else None
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class TestEmailConfigView(APIView):
    """
    Test email configuration (admin only for POST, public for GET)
    """
    permission_classes = [permissions.AllowAny]  # Allow anyone to check config
    
    def get(self, request):
        """Check if email is configured"""
        from django.conf import settings
        
        config_status = {
            'azure_client_id_set': bool(settings.AZURE_CLIENT_ID),
            'azure_client_secret_set': bool(settings.AZURE_CLIENT_SECRET),
            'azure_tenant_id_set': bool(settings.AZURE_TENANT_ID),
            'azure_sender_email': settings.AZURE_SENDER_EMAIL,
            'frontend_url': settings.FRONTEND_URL,
            'email_verification_hours': settings.EMAIL_VERIFICATION_TOKEN_EXPIRY_HOURS,
        }
        
        return Response(config_status)
    
    def post(self, request):
        """Send a test email - admin only"""
        # Check if user is admin
        if not (request.user and request.user.is_authenticated and request.user.designation == 'admin'):
            return Response(
                {'error': 'Admin access required'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        try:
            from utils.email_service import email_service
            
            test_email = request.data.get('email', request.user.email)
            
            html_content = """
            <h1>Test Email</h1>
            <p>This is a test email from the Task Management System.</p>
            <p>If you received this, email configuration is working correctly!</p>
            """
            
            success = email_service.send_email(
                to_email=test_email,
                subject='Test Email - Task Management System',
                html_content=html_content
            )
            
            if success:
                return Response({
                    'message': f'Test email sent successfully to {test_email}'
                }, status=status.HTTP_200_OK)
            else:
                return Response({
                    'error': 'Failed to send test email'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                
        except Exception as e:
            return Response({
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

